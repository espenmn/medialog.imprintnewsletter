# -*- coding: utf-8 -*-

from medialog.imprintnewsletter import _
from zope.interface import Interface
from zope.annotation.interfaces import IAnnotations
from Products.Five.browser import BrowserView
from persistent.list import PersistentList
from Products.statusmessages.interfaces import IStatusMessage
from zope.component import getMultiAdapter
from Products.Five.browser.pagetemplatefile import ViewPageTemplateFile
from Products.CMFPlone.utils import getSite
from email.utils import parseaddr 
from DateTime import DateTime

import io
# from datetime import datetime
# import xlsxwriter  # You need this package installed in your Plone/Python environment
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd


SUBSCRIBERS_KEY = 'medialog.imprint_newsletter.subscribers'
UNSUBSCRIBERS_KEY = 'medialog.imprint_newsletter.unsubscribers'

class IManageSubscribersView(Interface):
    """ Marker Interface for IManageSubscribersView"""

 

class SubscribeView(BrowserView):
    template = ViewPageTemplateFile('subscribe.pt')
    
    
    def redirect_view(self):
        return '/@@subscribe'

    def __call__(self):
        messages = IStatusMessage(self.request)
        if 'form.button_exportexcel' in self.request.form:
            return self.export_excel()
        elif 'form.button_importexcel' in self.request.form:
            fileobj = self.request.form['excel_file']
            if fileobj :
                return self.import_excel(fileobj)
            else:
                messages.add(f"No file or bad file.", type="error") 
        elif 'form.subscribed' in self.request.form:
            return self._handle_add()
        elif 'form.unsubscribe' in self.request.form:
            return self._handle_remove()
        return self.template()  # Calls template, avoids recursion
    
    def is_probably_email(self, s):
        name, addr = parseaddr(s)
        return '@' in addr and '.' in addr.split('@')[-1]

    def _get_unsubscribers(self):    
        site = getSite()
        annotations = IAnnotations(site)
        if UNSUBSCRIBERS_KEY not in annotations:
            annotations[UNSUBSCRIBERS_KEY] = PersistentList()
        return annotations[UNSUBSCRIBERS_KEY]

    def _get_subscribers(self):    
        site = getSite()
        annotations = IAnnotations(site)
        if SUBSCRIBERS_KEY not in annotations:
            annotations[SUBSCRIBERS_KEY] = PersistentList()
        return annotations[SUBSCRIBERS_KEY]

    
    def _handle_add(self):
        email = self.request.form.get('email', '').strip().lower()
        language = self.request.form.get('language') or None
        messages = IStatusMessage(self.request)
        if email and language:
            subscribers = self._get_subscribers()
            for email in email.split():
                exists = any(
                    item.get('email') == email and item.get('language') == language
                    for item in subscribers
                )

                if not exists and self.is_probably_email(email):
                    subscribers.append({
                        'email': email,
                        'language': language,
                        "created": DateTime()
                    })
                    messages.add("Successfully subscribed.", type="info")
                else:
                    messages.add("Already subscribed or not valid.", type="warning")
        else:
            messages.add("Email and language is required.", type="error")
        return self.request.response.redirect(self.context.absolute_url() + self.redirect_view() )

    def _handle_remove(self):
        raw_email = self.request.form.get('email', '')
        language = self.request.form.get('language') or None
        messages = IStatusMessage(self.request)
        subscribers = self._get_subscribers()
        unsubscribers = self._get_unsubscribers()

        emails = [e.strip().lower() for e in raw_email.split() if e.strip()]

        if not emails:
            messages.add("Email not found or invalid.", type="warning")
            return self.request.response.redirect(
                self.context.absolute_url() + self.redirect_view()
            )

        
        for email in emails:
            if not self.is_probably_email(email):
                messages.add(f"{email}: invalid email.", type="warning")
                continue

            # find matching entries
            matches = [item for item in subscribers if item.get('email', '').lower() == email]
            
            if not matches:
                messages.add(f"{email}: not found.", type="warning")
            
            # If they have subscribed in different languages
            if len(matches) > 1:
                matches = None
                if language:
                    matches = [ item for item in subscribers if item.get('email', '').lower() == email and item.get('language') == language  ]
                else:
                    messages.add(f"{email}: Yyou have subscribed in several languages. Please choose language", type="warning")
            
            if matches:
                for item in matches:
                    subscribers.remove(item)
                messages.add(f"{email}: unsubscribed successfully.", type="info")
                unsubscribers.append({
                        'email': email,
                        'language': language,
                         "created": DateTime()
                    })
            

        return self.request.response.redirect(
            self.context.absolute_url() + self.redirect_view()
        )
        
    def export_excel(self):
        # Create workbook and sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Subscribers"
        ws2 = wb.create_sheet("Unsubscribers")

        # Fill Subscribers sheet
        ws1.append(["E-mail", "Taal", "Aangemeld"])
        for sub in self.subscribers():
            ws1.append([sub["email"], sub["language"], sub["created"].strftime("%Y-%m-%d")])

        # Fill Unsubscribers sheet
        ws2.append(["E-mail", "Taal"])
        for unsub in self.unsubscribers():
            ws2.append([unsub["email"], unsub["language"], unsub["created"].strftime("%Y-%m-%d")])

        # Save workbook to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Return as HTTP response
        response = self.request.response
        response.setHeader(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.setHeader(
            "Content-Disposition",
            "attachment; filename=subscribers.xlsx"
        )
        response.write(output.read())
        return b""
    
    
    def import_excel(self, fileobj):    
        messages = IStatusMessage(self.request)
        # Ensure we have bytes in memory
        fileobj  = self.request.form['excel_file']
        file_bytes = fileobj.read()
        file_stream = io.BytesIO(file_bytes)
        
        
        # existing subscribers
        subscribers = self._get_subscribers()
        unsubscribers = self._get_unsubscribers()

        wb = load_workbook(filename=file_stream)
        ws = wb.active  # first sheet

        for row in ws.iter_rows(min_row=1, values_only=True):
            email = row[0]
            language = row[1] if row[1] else "nl"
            joined = row[2]

            # Default to today if missing
            if not joined:
                joined = DateTime()
            elif isinstance(joined, str):
                try:
                    joined = DateTime(joined)
                except Exception:
                    joined = DateTime()

            if email and self.is_probably_email(email):       
                exists = any(
                    item.get('email') == email  
                    for item in subscribers + unsubscribers
                ) 
                
                if not exists:        
                    subscribers.append({
                            'email': email,
                            'language': language,
                            "created": joined
                        })
                    messages.add(f"{email} successfully subscribed.", type="info")
                messages.add(f"{email} not added.", type="info")
                    
                    
                

        return self.template()
    
        
class ManageSubscribersView(SubscribeView):
    template = ViewPageTemplateFile('manage_subscribers.pt') 
    
    def redirect_view(self):
        return '/@@manage-subscribers'
    
    def subscribers(self):
        #TO DO, get this from utils
        data = self._get_subscribers()
        return sorted(data, key=lambda x: x['email']) 
    
    def unsubscribers(self):
        #TO DO, get this from utils
        data = self._get_unsubscribers()
        return sorted(data, key=lambda x: x['email']) 
    