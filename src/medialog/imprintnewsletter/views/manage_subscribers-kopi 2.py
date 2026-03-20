# -*- coding: utf-8 -*-

from medialog.imprintnewsletter import _
from zope.interface import Interface
from zope.annotation.interfaces import IAnnotations
from Products.Five.browser import BrowserView
from persistent.list import PersistentList
from Products.statusmessages.interfaces import IStatusMessage
from zope.component import getMultiAdapter
from Products.Five.browser.pagetemplatefile import ViewPageTemplateFile
from Products.CMFPlone.utils import getSite
from email.utils import parseaddr 
from DateTime import DateTime

import io
# from datetime import datetime
# import xlsxwriter  # You need this package installed in your Plone/Python environment
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd


SUBSCRIBERS_KEY = 'medialog.imprintnews_letter.subscribers'
UNSUBSCRIBERS_KEY = 'medialog.imprintnews_letter.unsubscribers'

class IManageSubscribersView(Interface):
    """ Marker Interface for IManageSubscribersView"""

 

class SubscribeView(BrowserView):
    template = ViewPageTemplateFile('subscribe.pt')
    
    
    def redirect_view(self):
        return '/@@subscribe'

    def __call__(self):
        messages = IStatusMessage(self.request)
        if 'form.button_exportexcel' in self.request.form:
            return self.export_excel()
        elif 'form.button_importexcel' in self.request.form:
            fileobj = self.request.form['excel_file']
            if fileobj :
                return self.import_excel(fileobj)
            else:
                messages.add(f"No file or bad file.", type="error") 
        elif 'form.subscribed' in self.request.form:
            return self._handle_add()
        elif 'form.unsubscribe' in self.request.form:
            return self._handle_remove()
        return self.template()  # Calls template, avoids recursion
    
    def is_probably_email(self, s):
        name, addr = parseaddr(s)
        return '@' in addr and '.' in addr.split('@')[-1]

    def _get_unsubscribers(self):    
        site = getSite()
        annotations = IAnnotations(site)
        if UNSUBSCRIBERS_KEY not in annotations:
            annotations[UNSUBSCRIBERS_KEY] = PersistentList()
        return annotations[UNSUBSCRIBERS_KEY]

    def _get_subscribers(self):    
        site = getSite()
        annotations = IAnnotations(site)
        if SUBSCRIBERS_KEY not in annotations:
            annotations[SUBSCRIBERS_KEY] = PersistentList()
        return annotations[SUBSCRIBERS_KEY]

    
    def _handle_add(self):
        email = self.request.form.get('email', '').strip().lower()
        language = self.request.form.get('language') or None
        messages = IStatusMessage(self.request)
        if email and language:
            subscribers = self._get_subscribers()
            for email in email.split():
                exists = any(
                    item.get('email') == email and item.get('language') == language
                    for item in subscribers
                )

                if not exists and self.is_probably_email(email):
                    subscribers.append({
                        'email': email,
                        'language': language,
                        "created": DateTime()
                    })
                    messages.add("Successfully subscribed.", type="info")
                else:
                    messages.add("Already subscribed or not valid.", type="warning")
        else:
            messages.add("Email and language is required.", type="error")
        return self.request.response.redirect(self.context.absolute_url() + self.redirect_view() )

    def _handle_remove(self):
        raw_email = self.request.form.get('email', '')
        language = self.request.form.get('language') or None
        messages = IStatusMessage(self.request)
        subscribers = self._get_subscribers()
        unsubscribers = self._get_unsubscribers()

        emails = [e.strip().lower() for e in raw_email.split() if e.strip()]

        if not emails:
            messages.add("Email not found or invalid.", type="warning")
            return self.request.response.redirect(
                self.context.absolute_url() + self.redirect_view()
            )

        
        for email in emails:
            if not self.is_probably_email(email):
                messages.add(f"{email}: invalid email.", type="warning")
                continue

            # find matching entries
            matches = [item for item in subscribers if item.get('email', '').lower() == email]
            
            if not matches:
                messages.add(f"{email}: not found.", type="warning")
            
            # If they have subscribed in different languages
            if len(matches) > 1:
                matches = None
                if language:
                    matches = [ item for item in subscribers if item.get('email', '').lower() == email and item.get('language') == language  ]
                else:
                    messages.add(f"{email}: You have subscribed in several languages. Please choose language", type="warning")
            
            if matches:
                for item in matches:
                    subscribers.remove(item)
                messages.add(f"{email}: unsubscribed successfully.", type="info")
                unsubscribers.append({
                        'email': email,
                        'language': language,
                        'created': DateTime()
                    })
            

        return self.request.response.redirect(
            self.context.absolute_url() + self.redirect_view()
        )
        
    def export_excel(self):
        # Create workbook and sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Subscribers"
        ws2 = wb.create_sheet("Unsubscribers")

        # Fill Subscribers sheet
        ws1.append(["E-mail", "Taal", "Aangemeld"])
        for sub in self.subscribers():
            ws1.append([sub["email"], sub["language"], sub["created"].strftime("%Y-%m-%d")])

        # Fill Unsubscribers sheet
        ws2.append(["E-mail", "Taal"])
        for unsub in self.unsubscribers():
            ws2.append([unsub["email"], unsub["language"], unsub["created"].strftime("%Y-%m-%d")])

        # Save workbook to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Return as HTTP response
        response = self.request.response
        response.setHeader(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.setHeader(
            "Content-Disposition",
            "attachment; filename=subscribers.xlsx"
        )
        response.write(output.read())
        return b""
    
    
    def import_excel(self, fileobj):    
        messages = IStatusMessage(self.request)

        # Get uploaded file
        fileobj = self.request.form.get('excel_file')
        if not fileobj or isinstance(fileobj, str):
            messages.add("No file uploaded", type="error")
            return

        # Read file into pandas
        file_bytes = fileobj.read()
        file_stream = io.BytesIO(file_bytes)
        try:
            df = pd.read_excel(file_stream)
        except Exception as e:
            messages.add(f"Error reading Excel: {e}", type="error")
            return

        # Normalize column names: lowercase, strip spaces
        df.columns = [str(c).strip().lower() for c in df.columns]
        
        # Map expected columns
        col_email = next((c for c in df.columns if "e-mail" in c), None)
        col_language = next((c for c in df.columns if "taal" in c), None)
        col_land = next((c for c in df.columns if "Naam land" in c), None)
        col_joined = next((c for c in df.columns if "aangemeld" in c), None)

        if not col_email:
            messages.add("No 'email' column found in Excel", type="error")
            return

        # Fill defaults
        if col_language is None:
            df["language"] = "nl"
            col_language = "language"
        else:
            df[col_language] = df[col_language].fillna("nl")

        if col_joined is None:
            df["joined"] = DateTime()
            col_joined = "joined"
        else:
            df[col_joined] = df[col_joined].apply(lambda x: DateTime(x) if pd.notnull(x) else DateTime())

        # Existing subscribers
        subscribers = self._get_subscribers()
        unsubscribers = self._get_unsubscribers()
        
        

        for _, row in df.iterrows():
            email = row[col_email]
            language = row[col_language]
            joined = row[col_joined]
            
            if email and self.is_probably_email(email):
                exists = any(item.get("email") == email for item in subscribers + unsubscribers)
                if not exists:
                    subscribers.append({
                        "email": email,
                        "language": language,
                        "created": joined
                    })
                    messages.add(f"{email}: successfully subscribed.", type="info")
                else:
                    messages.add(f"{email}: not added (already exists or is on unsubscribe list).", type="warning")    
            else:
                messages.add(f"{email}: not added (probably mailformed).", type="error")
        
        return self.template()
    
        
class ManageSubscribersView(SubscribeView):
    template = ViewPageTemplateFile('manage_subscribers.pt') 
    
    def redirect_view(self):
        return '/@@manage-subscribers'
    
    def subscribers(self):
        #TO DO, get this from utils
        data = self._get_subscribers()
        return sorted(data, key=lambda x: x['email']) 
    
    def unsubscribers(self):
        #TO DO, get this from utils
        data = self._get_unsubscribers()
        return sorted(data, key=lambda x: x['email']) 
    